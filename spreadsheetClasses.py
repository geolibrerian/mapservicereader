import csv
from openpyxl import Workbook



class SheetGenerate():
	
	def create(self, sheetname="NewSheet"):
		self.wb = Workbook()
		self.ws = self.wb.active
		self.ws.title = sheetname

	def newsheet(self,sheetname="NewSheet"):
		return self.wb.create_sheet(title=sheetname)

	def databycell(self,rows,cn=0,rn=0):
		for row in range(rn,len(rows)+1):
			for col in range(cn,len(rows[0])+1):
				self.ws.cell(column=col, row=row, value=rows[row][col])

	def databyrow(self,rows):
		[self.ws.append(row) for row in rows]


	def header(self,row):
		for col in range(0,len(row)+1):
			self.ws.cell(column=col, row=0, value=row[col])


	def save(self, filename):
		self.wb.save(filename)